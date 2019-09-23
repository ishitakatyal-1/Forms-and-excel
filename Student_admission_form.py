# -*- coding: utf-8 -*-
"""
Created on Sun Sep 22 16:58:06 2019

@author: Ishita
"""
import tkinter as tk
from tkinter import *
from openpyxl import *

wb = load_workbook('C:\\Users\\Ishita\\Desktop\\Python-GeeksForGeeks\\excel.xlsx')
sheet = wb.active

def excel_heading():
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 20
    sheet.cell(row=1, column=1).value = "Institutution ID"
    sheet.cell(row=1, column=3).value = "Name"
    sheet.cell(row=1, column=4).value = "Address"
    sheet.cell(row=1, column=4).value = "E-mail ID"
    sheet.cell(row=1, column=5).value = "Phone No."
    wb.save("C:\\Users\\Ishita\\Desktop\\Python-GeeksForGeeks\\excel.xlsx")
        
def clear():
    institution_id.delete(0, "end")
    name.delete(0, "end")
    address.delete(0, "end")
    email_id.delete(0, "end")
    phone.delete(0, "end")
    
def insert_record():
    if(institution_id.get() == "" or name.get() == "" or address.get() == "" or email_id.get() == "" or phone.get() == ""):
        print("Empty input not allowed")
    else:
        current_row = sheet.max_row
        sheet.cell(row=current_row + 1, column=1).value = institution_id.get()
        sheet.cell(row=current_row + 1, column=2).value = name.get()
        sheet.cell(row=current_row + 1, column=3).value = address.get()
        sheet.cell(row=current_row + 1, column=4).value = email_id.get()
        sheet.cell(row=current_row + 1, column=5).value = phone.get()
        #print("Records entered: ",sheet.max_row - 1)
        
        frame1 = tk.Tk()
        frame1.title("Records entered")
        
        frame2 = tk.Frame(frame1)
        frame2.pack()
         
        label00 = tk.Label(frame2, text=(sheet.max_row - 1), font="Times 20")
        label01 = tk.Label(frame2, text="Records entered", font="Times 20")
        button_esc = tk.Button(frame2, text="OK, Quit.", command= frame1.destroy)
        
        label00.grid(row=0, column=0, sticky=tk.W, pady=4)
        label01.grid(row=0, column=1, sticky=tk.W, pady=4)
        button_esc.grid(row=1, column=1, sticky=tk.W, pady=4)
        
        wb.save("C:\\Users\\Ishita\\Desktop\\Python-GeeksForGeeks\\excel.xlsx")
        clear()
        frame1.mainloop()

form = tk.Tk()
form.title("Registration Form")
form.geometry("500x500")
excel_heading()

mainframe1 = tk.Frame(form)
mainframe1.pack()


heading = tk.Label(mainframe1, text="Registration Form")
text1 = tk.Label(mainframe1, text="Insitution_ID")
institution_id = tk.Entry(mainframe1)
text2 = tk.Label(mainframe1, text="Name")
name = tk.Entry(mainframe1)
text3 = tk.Label(mainframe1, text="Address")
address = tk.Entry(mainframe1)
text4 = tk.Label(mainframe1, text="E-mail ID")
email_id = tk.Entry(mainframe1)
text5 = tk.Label(mainframe1, text="Phone No.")
phone = tk.Entry(mainframe1)
submit = tk.Button(mainframe1, text="Submit", command =insert_record)

heading.grid(row=0, column=0, columnspan=3, sticky=tk.W, pady=4)
text1.grid(row=1, column=0, columnspan=1, sticky=tk.W, pady=4)
institution_id.grid(row=1, column=1, columnspan=2, sticky=tk.W, pady=4)
text2.grid(row=2, column=0, columnspan=1, sticky=tk.W, pady=4)
name.grid(row=2, column=1, columnspan=2, sticky=tk.W, pady=4)
text3.grid(row=3, column=0, columnspan=1, sticky=tk.W, pady=4)
address.grid(row=3, column=1, columnspan=2, sticky=tk.W, pady=4)
text4.grid(row=4, column=0, columnspan=1, sticky=tk.W, pady=4)
email_id.grid(row=4, column=1, columnspan=2, sticky=tk.W, pady=4)
text5.grid(row=5, column=0, columnspan=1, sticky=tk.W, pady=4)
phone.grid(row=5, column=1, columnspan=2, sticky=tk.W, pady=4)
submit.grid(row=6, column=2, columnspan=1, sticky=tk.W, pady=4)

form.mainloop()        